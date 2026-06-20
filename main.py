# -*- coding: utf-8 -*-
"""
تطبيق الحسابات حسب الشُّعب
==========================
- إضافة الشُّعب (المجموعات)
- إضافة الأفراد داخل كل شعبة، ولكل فرد رصيد ورقم هاتف
- تحديد الكل / إضافة أو خصم مبلغ للجميع أو للمحدد أو لفرد
- نقل الفرد من شعبة إلى شعبة أخرى
- سجل حركات (معاملات) لكل فرد

الستاك: Kivy + KivyMD 1.1.1 + SQLite
العربية: arabic_reshaper للتشكيل + python-bidi للاتجاه على الأندرويد،
والخط يُسجَّل أيضًا باسم Roboto حتى تظهر عناوين الحقول بالعربية.
"""

import os
import sqlite3
from datetime import datetime

import arabic_reshaper

# على ويندوز إن ظهر عكس مزدوج، اجعلها False
USE_BIDI = True
try:
    from bidi.algorithm import get_display
except Exception:
    USE_BIDI = False

from kivy.lang import Builder
from kivy.clock import Clock
from kivy.metrics import dp
from kivy.properties import (
    StringProperty, NumericProperty, BooleanProperty, ObjectProperty
)
from kivy.uix.screenmanager import ScreenManager, Screen
from kivy.core.text import LabelBase
from kivy.core.window import Window
from kivy.properties import ColorProperty

from kivymd.app import MDApp
from kivymd.uix.boxlayout import MDBoxLayout
from kivymd.uix.card import MDCard
from kivymd.uix.label import MDLabel
from kivymd.uix.button import MDFlatButton, MDRaisedButton
from kivymd.uix.textfield import MDTextField
from kivymd.uix.dialog import MDDialog
from kivymd.uix.menu import MDDropdownMenu
from kivymd.uix.selectioncontrol import MDCheckbox
from kivymd.uix.snackbar import Snackbar
from kivy.base import ExceptionHandler, ExceptionManager
from kivy.logger import Logger


# ----------------------------------------------------------------------
# حماية ضد الخروج المفاجئ: أي خطأ غير متوقع في تنفيذ أمر يُسجَّل ويُتجاوز
# بدل أن يُغلق التطبيق نفسه.
# ----------------------------------------------------------------------
class _CrashGuard(ExceptionHandler):
    def handle_exception(self, exception):
        Logger.exception("Hesabat: تم تجاوز خطأ: %s", exception)
        return ExceptionManager.PASS


ExceptionManager.add_handler(_CrashGuard())


# ----------------------------------------------------------------------
# تسجيل خط عربي احترافي تلقائيًا من assets/fonts (وزن عادي + عريض)
# يُسجَّل باسمين: "Arabic" و"Roboto" (الأخير يجعل عناوين الحقول عربية)
# ----------------------------------------------------------------------
def _register_arabic_font():
    base = os.path.dirname(os.path.abspath(__file__))
    fonts_dir = os.path.join(base, "assets", "fonts")
    if not os.path.isdir(fonts_dir):
        return None
    ttfs = [os.path.join(fonts_dir, f) for f in os.listdir(fonts_dir)
            if f.lower().endswith(".ttf")]
    if not ttfs:
        return None
    regular = bold = None
    for p in ttfs:
        low = os.path.basename(p).lower()
        if "bold" in low or "semibold" in low:
            bold = p
        elif "regular" in low or "medium" in low:
            regular = p
    regular = regular or sorted(ttfs)[0]
    bold = bold or regular
    for name in ("Arabic", "Roboto"):
        try:
            LabelBase.register(name=name, fn_regular=regular, fn_bold=bold)
        except Exception:
            pass
    return "Arabic"


APP_FONT = _register_arabic_font()


# ----------------------------------------------------------------------
# نظام الألوان الاحترافي (كحلي/تيل عميق + رمادي ناعم)
# ----------------------------------------------------------------------
C_PRIMARY = (0.106, 0.286, 0.396, 1)    # #1B4965 كحلي احترافي
C_PRIMARY_DARK = (0.078, 0.227, 0.322, 1)  # #143A52
C_ACCENT = (0.231, 0.510, 0.553, 1)     # #3B828D تيل مهذّب
C_BG = (0.933, 0.945, 0.961, 1)         # #EEF1F5 خلفية رمادية ناعمة
C_SURFACE = (1, 1, 1, 1)                 # بطاقات بيضاء
C_SELECTED = (0.890, 0.925, 0.949, 1)   # #E3ECF2 تظليل التحديد
C_POSITIVE = (0.180, 0.490, 0.357, 1)   # #2E7D5B أخضر مهذّب
C_NEGATIVE = (0.753, 0.290, 0.239, 1)   # #C04A3D أحمر مهذّب


# ----------------------------------------------------------------------
# عرض النص العربي بشكل صحيح (تشكيل + اتجاه)
# ----------------------------------------------------------------------
def ar(text):
    if text is None:
        return ""
    reshaped = arabic_reshaper.reshape(str(text))
    if USE_BIDI:
        return get_display(reshaped)
    return reshaped


def phone_filter(substring, from_undo=False):
    """يسمح بالأرقام و + و المسافة فقط في حقل الهاتف."""
    return "".join(ch for ch in substring if ch in "0123456789+ ")


def _cell_str(v):
    """تحويل قيمة خلية إكسل إلى نص نظيف (مع معالجة الأرقام)."""
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _numstr(v):
    """تنسيق رقم لعرضه في حقل إدخال (بدون .0 الزائدة)."""
    if not v:
        return ""
    try:
        f = float(v)
        return str(int(f)) if f.is_integer() else str(f)
    except (ValueError, TypeError):
        return ""


def _download_dirs():
    """قائمة مجلدات محتملة للحفظ بالترتيب (أندرويد ثم سطح المكتب)."""
    dirs = []
    try:
        from android.storage import primary_external_storage_path
        dirs.append(os.path.join(primary_external_storage_path(), "Download"))
    except Exception:
        pass
    try:
        from android.storage import app_storage_path
        dirs.append(app_storage_path())
    except Exception:
        pass
    home = os.path.expanduser("~")
    dirs.append(os.path.join(home, "Downloads"))
    dirs.append(home)
    dirs.append(os.getcwd())
    return dirs


# ----------------------------------------------------------------------
# حقل إدخال عربي (RTL)
# يخزّن النص الخام في real_text ويعرضه مُشكّلًا بالاتجاه الصحيح،
# فتظهر الحروف متّصلة وغير معكوسة أثناء الكتابة.
# عند الحفظ تُقرأ القيمة من real_text وليس text.
# ----------------------------------------------------------------------
class ArabicInput(MDTextField):
    real_text = StringProperty("")

    def __init__(self, **kwargs):
        initial = kwargs.pop("text", "")
        super().__init__(**kwargs)
        self.halign = "right"
        if APP_FONT:
            self.font_name = APP_FONT
        if initial:
            self.real_text = initial
            self._render()

    def insert_text(self, substring, from_undo=False):
        self.real_text = self.real_text + substring
        self._render()

    def do_backspace(self, from_undo=False, mode="bkspc"):
        self.real_text = self.real_text[:-1]
        self._render()

    def _render(self):
        # نعرض النص مُشكّلًا دون المرور بـ insert_text (لا تكرار)
        self.text = ar(self.real_text) if self.real_text else ""
        try:
            self.cursor = (len(self.text), 0)
        except Exception:
            pass


# ----------------------------------------------------------------------
# طبقة قاعدة البيانات
# ----------------------------------------------------------------------
class DB:
    def __init__(self, path):
        self.conn = sqlite3.connect(path)
        self.conn.row_factory = sqlite3.Row
        self.conn.execute("PRAGMA foreign_keys = ON")
        self._create()

    def _create(self):
        c = self.conn.cursor()
        c.execute("""CREATE TABLE IF NOT EXISTS divisions(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            base_amount REAL DEFAULT 0,
            donation REAL DEFAULT 0)""")
        c.execute("""CREATE TABLE IF NOT EXISTS members(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            division_id INTEGER,
            balance REAL DEFAULT 0,
            phone TEXT DEFAULT '',
            mastercard TEXT DEFAULT '',
            FOREIGN KEY(division_id) REFERENCES divisions(id) ON DELETE CASCADE)""")
        c.execute("""CREATE TABLE IF NOT EXISTS transactions(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            member_id INTEGER,
            amount REAL,
            note TEXT,
            created_at TEXT,
            FOREIGN KEY(member_id) REFERENCES members(id) ON DELETE CASCADE)""")
        c.execute("""CREATE TABLE IF NOT EXISTS withdrawals(
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            amount REAL,
            note TEXT,
            created_at TEXT)""")
        # ترقية القواعد القديمة بإضv الأعمدة الجديدة إن لزم
        self._ensure_column(c, "members", "phone", "TEXT DEFAULT ''")
        self._ensure_column(c, "members", "mastercard", "TEXT DEFAULT ''")
        self._ensure_column(c, "divisions", "base_amount", "REAL DEFAULT 0")
        self._ensure_column(c, "divisions", "donation", "REAL DEFAULT 0")
        self.conn.commit()

    def _ensure_column(self, c, table, col, decl):
        cols = [r[1] for r in c.execute(f"PRAGMA table_info({table})").fetchall()]
        if col not in cols:
            c.execute(f"ALTER TABLE {table} ADD COLUMN {col} {decl}")

    # ----- الشُّعب -----
    def add_division(self, name):
        try:
            self.conn.execute("INSERT INTO divisions(name) VALUES(?)", (name,))
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False

    def update_division(self, did, name):
        self.conn.execute("UPDATE divisions SET name=? WHERE id=?", (name, did))
        self.conn.commit()

    def update_division_amounts(self, did, base_amount, donation):
        self.conn.execute(
            "UPDATE divisions SET base_amount=?, donation=? WHERE id=?",
            (base_amount, donation, did))
        self.conn.commit()

    def get_division(self, did):
        return self.conn.execute(
            "SELECT * FROM divisions WHERE id=?", (did,)).fetchone()

    def delete_division(self, did):
        self.conn.execute("DELETE FROM divisions WHERE id=?", (did,))
        self.conn.commit()

    def division_nathriya(self, did):
        """مجموع الخصومات (النثرية) لكل أفراد الشعبة كقيمة موجبة."""
        r = self.conn.execute(
            """SELECT COALESCE(SUM(-t.amount),0) AS p
               FROM transactions t JOIN members m ON t.member_id=m.id
               WHERE m.division_id=? AND t.amount<0""", (did,)).fetchone()
        return r["p"] if r else 0

    def get_divisions(self):
        q = """SELECT d.id, d.name, d.base_amount, d.donation,
                      COUNT(m.id) AS cnt,
                      COALESCE(SUM(m.balance),0) AS total,
                      COALESCE((SELECT SUM(-t.amount) FROM transactions t
                                JOIN members mm ON t.member_id=mm.id
                                WHERE mm.division_id=d.id AND t.amount<0),0)
                          AS nathriya
               FROM divisions d
               LEFT JOIN members m ON m.division_id = d.id
               GROUP BY d.id, d.name, d.base_amount, d.donation
               ORDER BY d.name"""
        return self.conn.execute(q).fetchall()

    # ----- الأفراد -----
    def add_member(self, name, division_id, phone="", mastercard=""):
        self.conn.execute(
            "INSERT INTO members(name, division_id, balance, phone, mastercard) "
            "VALUES(?,?,0,?,?)", (name, division_id, phone, mastercard))
        self.conn.commit()

    def update_member(self, mid, name, phone, mastercard=""):
        self.conn.execute(
            "UPDATE members SET name=?, phone=?, mastercard=? WHERE id=?",
            (name, phone, mastercard, mid))
        self.conn.commit()

    def delete_member(self, mid):
        self.conn.execute("DELETE FROM members WHERE id=?", (mid,))
        self.conn.commit()

    def move_member(self, mid, new_division_id):
        self.conn.execute("UPDATE members SET division_id=? WHERE id=?",
                          (new_division_id, mid))
        self.conn.commit()

    def get_members(self, division_id):
        return self.conn.execute(
            "SELECT * FROM members WHERE division_id=? ORDER BY name",
            (division_id,)).fetchall()

    def get_member(self, mid):
        return self.conn.execute(
            "SELECT * FROM members WHERE id=?", (mid,)).fetchone()

    # ----- المعاملات -----
    def add_transaction(self, member_id, amount, note=""):
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        self.conn.execute(
            "INSERT INTO transactions(member_id, amount, note, created_at) "
            "VALUES(?,?,?,?)", (member_id, amount, note, now))
        self.conn.execute(
            "UPDATE members SET balance = balance + ? WHERE id=?",
            (amount, member_id))
        self.conn.commit()

    def bulk_add(self, member_ids, amount, note=""):
        for mid in member_ids:
            self.add_transaction(mid, amount, note)

    def get_transactions(self, member_id):
        return self.conn.execute(
            "SELECT * FROM transactions WHERE member_id=? ORDER BY id DESC",
            (member_id,)).fetchall()

    # ----- النثرية العامة (صندوق واحد مجموع من كل الشُّعب) -----
    def nathriya_collected(self):
        """مجموع كل الخصومات في كل الشُّعب (يُجمع في النثرية)."""
        r = self.conn.execute(
            "SELECT COALESCE(SUM(-amount),0) AS p FROM transactions "
            "WHERE amount<0").fetchone()
        return r["p"] if r else 0

    def nathriya_withdrawn(self):
        r = self.conn.execute(
            "SELECT COALESCE(SUM(amount),0) AS p FROM withdrawals").fetchone()
        return r["p"] if r else 0

    def nathriya_available(self):
        return self.nathriya_collected() - self.nathriya_withdrawn()

    def add_withdrawal(self, amount, note=""):
        now = datetime.now().strftime("%Y-%m-%d %H:%M")
        self.conn.execute(
            "INSERT INTO withdrawals(amount, note, created_at) VALUES(?,?,?)",
            (amount, note, now))
        self.conn.commit()


# ----------------------------------------------------------------------
# عناصر القوائم
# ----------------------------------------------------------------------
class DivisionCard(MDCard):
    div_id = NumericProperty(0)
    name_ar = StringProperty("")
    count_ar = StringProperty("")
    base_ar = StringProperty("")
    nath_ar = StringProperty("")
    dona_ar = StringProperty("")
    net_ar = StringProperty("")


class MemberCard(MDCard):
    member_id = NumericProperty(0)
    name_ar = StringProperty("")
    balance_ar = StringProperty("")
    phone_ar = StringProperty("")
    selected = BooleanProperty(False)
    selecting = BooleanProperty(False)


class TxRow(MDBoxLayout):
    amount_ar = StringProperty("")
    note_ar = StringProperty("")
    date_str = StringProperty("")
    is_add = BooleanProperty(True)


# ----------------------------------------------------------------------
# الشاشات
# ----------------------------------------------------------------------
class DivisionsScreen(Screen):
    def on_pre_enter(self, *a):
        Clock.schedule_once(lambda *_: self.refresh(), 0)

    def refresh(self):
        if "div_list" not in self.ids:
            return
        app = MDApp.get_running_app()
        avail = app.db.nathriya_available()
        if "toolbar" in self.ids:
            self.ids.toolbar.title = ar(f"الصندوق   |   النثرية: {avail:.0f}")
        box = self.ids.div_list
        box.clear_widgets()
        for r in app.db.get_divisions():
            net = (r["base_amount"] or 0) - (r["nathriya"] or 0) - (r["donation"] or 0)
            card = DivisionCard()
            card.div_id = r["id"]
            card.name_ar = ar(r["name"])
            card.count_ar = ar(f"الأفراد: {r['cnt']}")
            card.base_ar = ar(f"المبلغ الأصلي: {r['base_amount']:.0f}")
            card.nath_ar = ar(f"النثرية: {r['nathriya']:.0f}")
            card.dona_ar = ar(f"التبرع: {r['donation']:.0f}")
            card.net_ar = ar(f"الصافي: {net:.0f}")
            box.add_widget(card)

    def withdraw_nathriya(self):
        app = MDApp.get_running_app()
        avail = app.db.nathriya_available()
        app.amount_dialog(
            title=f"سحب من النثرية (المتاح: {avail:.0f})",
            show_sign=False,
            on_ok=lambda amount, note: (
                app.db.add_withdrawal(amount, note),
                self.refresh(),
                toast("تم السحب من النثرية")))

    def open_add_dialog(self):
        MDApp.get_running_app().text_dialog(
            title="إضافة شعبة", hint="اسم الشعبة", on_ok=self._do_add)

    def _do_add(self, value):
        value = value.strip()
        if not value:
            return
        if not MDApp.get_running_app().db.add_division(value):
            toast("اسم الشعبة موجود مسبقًا")
        self.refresh()

    def edit_division(self, div_id, current_name):
        app = MDApp.get_running_app()
        app.text_dialog(
            title="تعديل الشعبة", hint="اسم الشعبة", text=current_name,
            on_ok=lambda v: (app.db.update_division(div_id, v.strip()),
                             self.refresh()))

    def edit_amounts(self, div_id):
        app = MDApp.get_running_app()
        d = app.db.get_division(div_id)
        if not d:
            return
        app.division_amounts_dialog(
            base=d["base_amount"], donation=d["donation"],
            on_ok=lambda b, dn: (app.db.update_division_amounts(div_id, b, dn),
                                 self.refresh()))

    def delete_division(self, div_id):
        app = MDApp.get_running_app()
        app.confirm_dialog(
            "حذف الشعبة وكل أفرادها؟",
            lambda: (app.db.delete_division(div_id), self.refresh()))

    def open_division(self, div_id, name):
        ms = self.manager.get_screen("members")
        ms.division_id = div_id
        ms.division_name = name
        self.manager.current = "members"

    # ----- استيراد من إكسل -----
    def import_excel(self):
        try:
            from plyer import filechooser
        except Exception:
            toast("منتقي الملفات غير متوفر")
            return
        try:
            filechooser.open_file(
                title="اختر ملف إكسل",
                filters=[["Excel", "*.xlsx", "*.xls"]],
                on_selection=self._on_excel)
        except Exception:
            toast("تعذّر فتح منتقي الملفات")

    def _on_excel(self, selection):
        if not selection:
            return
        path = selection[0] if isinstance(selection, (list, tuple)) else selection
        Clock.schedule_once(lambda dt: self._after_pick(path), 0)

    def _after_pick(self, path):
        added, skipped = self._import_excel_from_path(path)
        self.refresh()
        toast(f"تمت إضافة {added} فرد" + (f" (تخطّي {skipped})" if skipped else ""))

    def _import_excel_from_path(self, path):
        """يقرأ ملف إكسل بالأعمدة: الاسم | الهاتف | الشعبة ويضيف الأفراد."""
        try:
            import openpyxl
        except Exception:
            toast("مكتبة openpyxl غير مثبتة")
            return (0, 0)
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            toast("تعذّر فتح الملف")
            return (0, 0)
        ws = wb.active
        db = MDApp.get_running_app().db
        headers = {"الاسم", "اسم", "اسم الفرد", "name", "الاسم الكامل"}
        added = skipped = 0
        for idx, row in enumerate(ws.iter_rows(values_only=True)):
            if not row:
                continue
            name = _cell_str(row[0] if len(row) > 0 else "")
            phone = _cell_str(row[1] if len(row) > 1 else "")
            div = _cell_str(row[2] if len(row) > 2 else "")
            mastercard = _cell_str(row[3] if len(row) > 3 else "")
            if not name:
                continue
            if idx == 0 and name in headers:   # تخطّي سطر العناوين
                continue
            if not div:
                div = "غير مصنّف"
            db.add_division(div)               # يتجاهل المكرر
            did = next((d["id"] for d in db.get_divisions()
                        if d["name"] == div), None)
            if did is None:
                skipped += 1
                continue
            db.add_member(name, did, phone, mastercard)
            added += 1
        try:
            wb.close()
        except Exception:
            pass
        return (added, skipped)


class MembersScreen(Screen):
    division_id = NumericProperty(0)
    division_name = StringProperty("")
    selecting = BooleanProperty(False)

    def on_pre_enter(self, *a):
        self.selecting = False
        Clock.schedule_once(self._enter, 0)

    def _enter(self, *a):
        # احفظ شريط التحديد وأخرجه من العرض حتى يُطلب (يمنع ظهور أي ظل)
        if not hasattr(self, "_sel_bar"):
            self._sel_bar = self.ids.sel_bar
            self._col = self.ids.col
        if self._sel_bar.parent:
            self._col.remove_widget(self._sel_bar)
        self.refresh()

    def refresh(self):
        if "member_list" not in self.ids:
            return
        self.ids.toolbar.title = ar(self.division_name)
        box = self.ids.member_list
        box.clear_widgets()
        for r in MDApp.get_running_app().db.get_members(self.division_id):
            card = MemberCard()
            card.member_id = r["id"]
            card.name_ar = ar(r["name"])
            card.balance_ar = ar(f"الرصيد: {r['balance']:.0f}")
            card.phone_ar = r["phone"] or ""
            card.selecting = self.selecting
            box.add_widget(card)

    def toggle_select_mode(self):
        if not hasattr(self, "_sel_bar"):
            self._sel_bar = self.ids.sel_bar
            self._col = self.ids.col
            if self._sel_bar.parent:
                self._col.remove_widget(self._sel_bar)
        self.selecting = not self.selecting
        if self.selecting and not self._sel_bar.parent:
            self._col.add_widget(self._sel_bar, index=1)
        elif not self.selecting and self._sel_bar.parent:
            self._col.remove_widget(self._sel_bar)
        for w in self.ids.member_list.children:
            w.selecting = self.selecting
            if not self.selecting:
                w.selected = False

    def select_all(self, value=True):
        for w in self.ids.member_list.children:
            w.selected = value

    def _selected_ids(self):
        return [w.member_id for w in self.ids.member_list.children if w.selected]

    def open_add_member(self):
        MDApp.get_running_app().member_dialog(
            "إضافة فرد", on_ok=self._do_add_member)

    def _do_add_member(self, name, phone, mastercard=""):
        MDApp.get_running_app().db.add_member(
            name, self.division_id, phone, mastercard)
        self.refresh()

    def apply_all(self):
        ids = [r["id"] for r in
               MDApp.get_running_app().db.get_members(self.division_id)]
        if not ids:
            toast("لا يوجد أفراد")
            return
        self._amount(ids, "إضافة / خصم لكل الأفراد")

    def apply_selected(self):
        ids = self._selected_ids()
        if not ids:
            toast("لم تحدد أي فرد")
            return
        self._amount(ids, f"إضافة / خصم لـ {len(ids)} فرد")

    def _amount(self, ids, title):
        app = MDApp.get_running_app()
        app.amount_dialog(
            title=title, show_sign=True,
            on_ok=lambda amount, note: (
                app.db.bulk_add(ids, amount, note),
                self.refresh(),
                toast("تم تطبيق العملية")))

    def open_member(self, mid):
        if self.selecting:
            return
        ms = self.manager.get_screen("member")
        ms.member_id = mid
        ms.back_division = self.division_id
        self.manager.current = "member"

    # ----- استيراد أفراد لهذه الشعبة -----
    def import_excel(self):
        try:
            from plyer import filechooser
        except Exception:
            toast("منتقي الملفات غير متوفر")
            return
        try:
            filechooser.open_file(
                title="اختر ملف إكسل",
                filters=[["Excel", "*.xlsx", "*.xls"]],
                on_selection=self._on_import)
        except Exception:
            toast("تعذّر فتح منتقي الملفات")

    def _on_import(self, selection):
        if not selection:
            return
        path = selection[0] if isinstance(selection, (list, tuple)) else selection
        Clock.schedule_once(lambda dt: self._do_import(path), 0)

    def _do_import(self, path):
        added = self._import_members_here(path)
        self.refresh()
        toast(f"تمت إضافة {added} فرد للشعبة")

    def _import_members_here(self, path):
        """يقرأ: الاسم | الهاتف | ماستر كارد، ويضيف للشعبة الحالية.
        يدعم أيضًا قالب 4 أعمدة (الاسم|الهاتف|الشعبة|ماستر كارد)."""
        try:
            import openpyxl
        except Exception:
            toast("مكتبة openpyxl غير مثبتة")
            return 0
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            toast("تعذّر فتح الملف")
            return 0
        db = MDApp.get_running_app().db
        rows = list(wb.active.iter_rows(values_only=True))
        headers = {"الاسم", "اسم", "اسم الفرد", "name"}
        mc_col, start = 2, 0
        if rows:
            h0 = _cell_str(rows[0][0] if len(rows[0]) > 0 else "")
            h2 = _cell_str(rows[0][2] if len(rows[0]) > 2 else "")
            if h0 in headers:
                start = 1
                if "شعب" in h2 or "division" in h2.lower():
                    mc_col = 3
        added = 0
        for row in rows[start:]:
            if not row:
                continue
            name = _cell_str(row[0] if len(row) > 0 else "")
            phone = _cell_str(row[1] if len(row) > 1 else "")
            mc = _cell_str(row[mc_col] if len(row) > mc_col else "")
            if not name:
                continue
            db.add_member(name, self.division_id, phone, mc)
            added += 1
        try:
            wb.close()
        except Exception:
            pass
        return added

    # ----- تصدير أفراد هذه الشعبة إلى إكسل -----
    def export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except Exception:
            toast("مكتبة openpyxl غير مثبتة")
            return
        db = MDApp.get_running_app().db
        members = db.get_members(self.division_id)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "الأفراد"
        try:
            ws.sheet_view.rightToLeft = True
        except Exception:
            pass
        ws.append(["الاسم", "الهاتف", "ماستر كارد", "الرصيد"])
        for m in members:
            ws.append([m["name"], m["phone"] or "", m["mastercard"] or "",
                       m["balance"]])
        try:
            fill = PatternFill("solid", fgColor="1B4965")
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = fill
                cell.alignment = Alignment(horizontal="center")
            for col, w in zip("ABCD", (26, 16, 22, 14)):
                ws.column_dimensions[col].width = w
        except Exception:
            pass
        safe = "".join(ch for ch in (self.division_name or "شعبة")
                       if ch not in '\\/:*?"<>|').strip() or "شعبة"
        fname = f"{safe}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        path = self._save_workbook(wb, fname)
        if path:
            MDApp.get_running_app().show_path_dialog(path)
        else:
            toast("تعذّر حفظ الملف")

    def _save_workbook(self, wb, fname):
        for d in _download_dirs():
            try:
                os.makedirs(d, exist_ok=True)
                p = os.path.join(d, fname)
                wb.save(p)
                return p
            except Exception:
                continue
        return None


class MemberScreen(Screen):
    member_id = NumericProperty(0)
    back_division = NumericProperty(0)
    member_name = StringProperty("")
    phone_val = StringProperty("")
    mc_val = StringProperty("")
    move_menu = ObjectProperty(None, allownone=True)

    def on_pre_enter(self, *a):
        Clock.schedule_once(lambda *_: self.refresh(), 0)

    def refresh(self):
        if "tx_list" not in self.ids:
            return
        app = MDApp.get_running_app()
        m = app.db.get_member(self.member_id)
        if not m:
            self.manager.current = "members"
            return
        self.member_name = m["name"]
        self.phone_val = m["phone"] or ""
        self.mc_val = m["mastercard"] or ""
        self.ids.toolbar.title = ar(m["name"])
        self.ids.balance_lbl.text = ar(f"الرصيد الحالي: {m['balance']:.0f}")
        self.ids.phone_lbl.text = self.phone_val or ar("لا يوجد رقم هاتف")
        self.ids.mc_lbl.text = self.mc_val or ar("لا يوجد ماستر كارد")

        box = self.ids.tx_list
        box.clear_widgets()
        for t in app.db.get_transactions(self.member_id):
            row = TxRow()
            row.is_add = t["amount"] >= 0
            sign = "+" if t["amount"] >= 0 else "-"
            row.amount_ar = f"{sign}{abs(t['amount']):.0f}"
            row.note_ar = ar(t["note"] or "—")
            row.date_str = t["created_at"]
            box.add_widget(row)

    def copy_text(self, value, label):
        if not value or not value.strip():
            toast("لا يوجد ما يُنسخ")
            return
        try:
            from kivy.core.clipboard import Clipboard
            Clipboard.copy(value.strip())
            toast(f"تم نسخ {label}")
        except Exception:
            toast("تعذّر النسخ")

    def add_amount(self, positive=True):
        app = MDApp.get_running_app()
        title = "إضافة مبلغ" if positive else "خصم مبلغ"
        app.amount_dialog(
            title=title, show_sign=False,
            on_ok=lambda amount, note: (
                app.db.add_transaction(
                    self.member_id, amount if positive else -amount, note),
                self.refresh()))

    def open_move_menu(self, caller):
        app = MDApp.get_running_app()
        items = []
        for d in app.db.get_divisions():
            if d["id"] == self.back_division:
                continue
            items.append({
                "viewclass": "OneLineListItem",
                "text": ar(d["name"]),
                "height": dp(48),
                "on_release": (lambda did=d["id"]: self._do_move(did)),
            })
        if not items:
            toast("لا توجد شعبة أخرى")
            return
        self.move_menu = MDDropdownMenu(caller=caller, items=items, width_mult=4)
        self.move_menu.open()

    def _do_move(self, new_div):
        if self.move_menu:
            self.move_menu.dismiss()
        MDApp.get_running_app().db.move_member(self.member_id, new_div)
        toast("تم نقل الفرد")
        self.manager.current = "members"

    def edit_member(self):
        app = MDApp.get_running_app()
        m = app.db.get_member(self.member_id)
        if not m:
            return
        app.member_dialog(
            "تعديل الفرد", name=m["name"], phone=(m["phone"] or ""),
            mastercard=(m["mastercard"] or ""),
            on_ok=lambda nm, ph, mc: (
                app.db.update_member(self.member_id, nm, ph, mc),
                self.refresh()))

    def delete(self):
        app = MDApp.get_running_app()
        app.confirm_dialog(
            "حذف هذا الفرد؟",
            lambda: (app.db.delete_member(self.member_id),
                     setattr(self.manager, "current", "members")))


# ----------------------------------------------------------------------
def toast(msg):
    try:
        Snackbar(text=ar(msg)).open()
    except Exception:
        print(msg)


# ----------------------------------------------------------------------
# واجهة KV
# ----------------------------------------------------------------------
KV = '''
#:import dp kivy.metrics.dp

<DivisionCard>:
    orientation: "vertical"
    size_hint_y: None
    height: dp(174)
    padding: dp(14)
    spacing: dp(2)
    radius: [dp(16)]
    elevation: 1
    md_bg_color: app.c_surface
    ripple_behavior: True
    on_release:
        app.root.get_screen("divisions").open_division(root.div_id, root.name_ar)
    MDBoxLayout:
        size_hint_y: None
        height: dp(36)
        MDLabel:
            text: root.name_ar
            halign: "right"
            font_style: "H6"
            bold: True
        MDIconButton:
            icon: "cash-multiple"
            theme_text_color: "Custom"
            text_color: app.c_accent
            on_release:
                app.root.get_screen("divisions").edit_amounts(root.div_id)
        MDIconButton:
            icon: "pencil"
            on_release:
                app.root.get_screen("divisions").edit_division(root.div_id, root.name_ar)
        MDIconButton:
            icon: "delete"
            theme_text_color: "Custom"
            text_color: app.c_negative
            on_release:
                app.root.get_screen("divisions").delete_division(root.div_id)
    MDLabel:
        text: root.count_ar
        halign: "right"
        font_style: "Caption"
        theme_text_color: "Secondary"
        size_hint_y: None
        height: dp(20)
    MDBoxLayout:
        size_hint_y: None
        height: dp(22)
        MDLabel:
            text: root.base_ar
            halign: "right"
            font_style: "Caption"
        MDLabel:
            text: root.nath_ar
            halign: "right"
            font_style: "Caption"
            theme_text_color: "Custom"
            text_color: app.c_negative
    MDBoxLayout:
        size_hint_y: None
        height: dp(26)
        MDLabel:
            text: root.dona_ar
            halign: "right"
            font_style: "Caption"
            theme_text_color: "Custom"
            text_color: app.c_negative
        MDLabel:
            text: root.net_ar
            halign: "right"
            font_style: "Subtitle2"
            bold: True
            theme_text_color: "Custom"
            text_color: app.c_primary

<MemberCard>:
    size_hint_y: None
    height: dp(84)
    padding: dp(12)
    spacing: dp(8)
    radius: [dp(16)]
    elevation: 1
    md_bg_color: app.c_selected if root.selected else app.c_surface
    ripple_behavior: True
    on_release:
        app.root.get_screen("members").open_member(root.member_id) if not root.selecting else setattr(root, "selected", not root.selected)
    MDCheckbox:
        size_hint_x: None
        width: dp(36) if root.selecting else 0
        opacity: 1 if root.selecting else 0
        disabled: not root.selecting
        active: root.selected
        on_active: root.selected = self.active
    MDBoxLayout:
        orientation: "vertical"
        MDLabel:
            text: root.name_ar
            halign: "right"
            font_style: "Subtitle1"
            bold: True
        MDLabel:
            text: root.balance_ar
            halign: "right"
            font_style: "Caption"
            theme_text_color: "Secondary"
        MDLabel:
            text: root.phone_ar
            halign: "right"
            font_style: "Caption"
            theme_text_color: "Hint"
            size_hint_y: None
            height: self.texture_size[1] if root.phone_ar else 0
            opacity: 1 if root.phone_ar else 0

<TxRow>:
    size_hint_y: None
    height: dp(56)
    padding: dp(10), dp(4)
    spacing: dp(8)
    MDLabel:
        text: root.amount_ar
        halign: "left"
        size_hint_x: None
        width: dp(90)
        bold: True
        font_style: "Subtitle1"
        theme_text_color: "Custom"
        text_color: app.c_positive if root.is_add else app.c_negative
    MDBoxLayout:
        orientation: "vertical"
        MDLabel:
            text: root.note_ar
            halign: "right"
            font_style: "Body2"
        MDLabel:
            text: root.date_str
            halign: "right"
            font_style: "Caption"
            theme_text_color: "Hint"

ScreenManager:
    DivisionsScreen:
        name: "divisions"
    MembersScreen:
        name: "members"
    MemberScreen:
        name: "member"

<DivisionsScreen>:
    MDBoxLayout:
        orientation: "vertical"
        MDTopAppBar:
            id: toolbar
            title: app.t_app
            elevation: 3
            md_bg_color: app.c_primary
            specific_text_color: 1, 1, 1, 1
            right_action_items: [["cash-minus", lambda x: root.withdraw_nathriya()], ["microsoft-excel", lambda x: root.import_excel()], ["plus", lambda x: root.open_add_dialog()]]
        ScrollView:
            MDBoxLayout:
                id: div_list
                orientation: "vertical"
                spacing: dp(10)
                padding: dp(12)
                adaptive_height: True
        MDLabel:
            text: app.t_credit
            halign: "center"
            font_style: "Caption"
            theme_text_color: "Hint"
            size_hint_y: None
            height: dp(28)

<MembersScreen>:
    MDBoxLayout:
        id: col
        orientation: "vertical"
        MDTopAppBar:
            id: toolbar
            title: ""
            elevation: 3
            md_bg_color: app.c_primary
            specific_text_color: 1, 1, 1, 1
            left_action_items: [["arrow-right", lambda x: setattr(root.manager, "current", "divisions")]]
            right_action_items: [["file-export", lambda x: root.export_excel()], ["file-import", lambda x: root.import_excel()], ["check-all", lambda x: root.toggle_select_mode()]]
        ScrollView:
            MDBoxLayout:
                id: member_list
                orientation: "vertical"
                spacing: dp(8)
                padding: dp(12)
                adaptive_height: True
        MDBoxLayout:
            id: sel_bar
            size_hint_y: None
            height: dp(56)
            padding: dp(8)
            spacing: dp(8)
            MDRaisedButton:
                text: app.t_select_all
                md_bg_color: app.c_primary
                on_release: root.select_all(True)
            MDRaisedButton:
                text: app.t_apply_sel
                md_bg_color: app.c_positive
                on_release: root.apply_selected()
        MDBoxLayout:
            size_hint_y: None
            height: dp(56)
            padding: dp(8)
            spacing: dp(8)
            MDRaisedButton:
                text: app.t_add_member
                md_bg_color: app.c_primary
                on_release: root.open_add_member()
            MDRaisedButton:
                text: app.t_apply_all
                md_bg_color: app.c_accent
                on_release: root.apply_all()

<MemberScreen>:
    MDBoxLayout:
        orientation: "vertical"
        MDTopAppBar:
            id: toolbar
            title: ""
            elevation: 3
            md_bg_color: app.c_primary
            specific_text_color: 1, 1, 1, 1
            left_action_items: [["arrow-right", lambda x: setattr(root.manager, "current", "members")]]
            right_action_items: [["account-arrow-right", lambda x: root.open_move_menu(x)], ["pencil", lambda x: root.edit_member()], ["delete", lambda x: root.delete()]]
        MDCard:
            size_hint_y: None
            height: dp(80)
            padding: dp(16)
            radius: [dp(16)]
            elevation: 2
            md_bg_color: app.c_primary
            MDLabel:
                id: balance_lbl
                text: ""
                halign: "center"
                font_style: "H6"
                bold: True
                theme_text_color: "Custom"
                text_color: 1, 1, 1, 1
        MDBoxLayout:
            size_hint_y: None
            height: dp(40)
            padding: dp(14), 0
            spacing: dp(4)
            MDIcon:
                icon: "phone"
                size_hint_x: None
                width: dp(24)
                theme_text_color: "Secondary"
            MDLabel:
                id: phone_lbl
                text: ""
                halign: "left"
                font_style: "Body2"
                theme_text_color: "Secondary"
            MDIconButton:
                icon: "content-copy"
                theme_text_color: "Custom"
                text_color: app.c_accent
                on_release: root.copy_text(root.phone_val, "رقم الهاتف")
        MDBoxLayout:
            size_hint_y: None
            height: dp(40)
            padding: dp(14), 0
            spacing: dp(4)
            MDIcon:
                icon: "credit-card"
                size_hint_x: None
                width: dp(24)
                theme_text_color: "Secondary"
            MDLabel:
                id: mc_lbl
                text: ""
                halign: "left"
                font_style: "Body2"
                theme_text_color: "Secondary"
            MDIconButton:
                icon: "content-copy"
                theme_text_color: "Custom"
                text_color: app.c_accent
                on_release: root.copy_text(root.mc_val, "رقم الماستر كارد")
        MDBoxLayout:
            size_hint_y: None
            height: dp(56)
            padding: dp(8)
            spacing: dp(8)
            MDRaisedButton:
                text: app.t_add_amount
                md_bg_color: app.c_positive
                on_release: root.add_amount(True)
            MDRaisedButton:
                text: app.t_deduct_amount
                md_bg_color: app.c_negative
                on_release: root.add_amount(False)
        ScrollView:
            MDBoxLayout:
                id: tx_list
                orientation: "vertical"
                spacing: dp(4)
                padding: dp(8)
                adaptive_height: True
'''


# ----------------------------------------------------------------------
# التطبيق
# ----------------------------------------------------------------------
class HesabatApp(MDApp):
    # نصوص الواجهة الثابتة (مُشكَّلة مسبقًا لتظهر الحروف متصلة)
    t_app = StringProperty(ar("الصندوق"))
    t_credit = StringProperty(ar("تطوير عمر كريم"))
    t_add_member = StringProperty(ar("إضافة فرد"))
    t_apply_all = StringProperty(ar("إضافة / خصم للكل"))
    t_apply_sel = StringProperty(ar("إضافة / خصم للمحدد"))
    t_select_all = StringProperty(ar("تحديد الكل"))
    t_add_amount = StringProperty(ar("إضافة مبلغ"))
    t_deduct_amount = StringProperty(ar("خصم مبلغ"))

    # ألوان الواجهة (قابلة للتعديل من مكان واحد)
    c_primary = ColorProperty(C_PRIMARY)
    c_accent = ColorProperty(C_ACCENT)
    c_surface = ColorProperty(C_SURFACE)
    c_selected = ColorProperty(C_SELECTED)
    c_positive = ColorProperty(C_POSITIVE)
    c_negative = ColorProperty(C_NEGATIVE)

    def build(self):
        self.theme_cls.theme_style = "Light"
        self.theme_cls.primary_palette = "BlueGray"
        self.theme_cls.primary_hue = "800"
        self.theme_cls.accent_palette = "Teal"
        self.theme_cls.material_style = "M2"
        Window.clearcolor = C_BG
        try:
            from android.permissions import request_permissions, Permission
            request_permissions([Permission.WRITE_EXTERNAL_STORAGE,
                                 Permission.READ_EXTERNAL_STORAGE])
        except Exception:
            pass
        _icon = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                             "assets", "icon.png")
        if os.path.exists(_icon):
            self.icon = _icon
        if APP_FONT:
            for style in self.theme_cls.font_styles.keys():
                if style == "Icon":
                    continue
                self.theme_cls.font_styles[style][0] = APP_FONT
        db_path = os.path.join(self.user_data_dir, "hesabat.db") \
            if hasattr(self, "user_data_dir") else "hesabat.db"
        self.db = DB(db_path)
        return Builder.load_string(KV)

    # ---- حوار إدخال نصّي (للشُّعب) ----
    def text_dialog(self, title, hint, text="", on_ok=None):
        field = ArabicInput(hint_text=ar(hint), text=text)
        dlg = MDDialog(
            title=ar(title), type="custom", content_cls=field,
            buttons=[
                MDFlatButton(text=ar("إلغاء"), on_release=lambda x: dlg.dismiss()),
                MDRaisedButton(
                    text=ar("حفظ"),
                    on_release=lambda x: (on_ok and on_ok(field.real_text),
                                          dlg.dismiss())),
            ])
        dlg.open()

    # ---- حوار إضافة/تعديل فرد (اسم + هاتف + ماستر كارد) ----
    def member_dialog(self, title, name="", phone="", mastercard="", on_ok=None):
        name_field = ArabicInput(hint_text=ar("اسم الفرد"), text=name)
        phone_field = MDTextField(
            hint_text=ar("رقم الهاتف"), text=phone,
            halign="left", input_filter=phone_filter)
        mc_field = MDTextField(
            hint_text=ar("ماستر كارد"), text=mastercard,
            halign="left", input_filter=phone_filter)
        box = MDBoxLayout(
            orientation="vertical", spacing=dp(8),
            size_hint_y=None, height=dp(210))
        box.add_widget(name_field)
        box.add_widget(phone_field)
        box.add_widget(mc_field)

        def _ok(*a):
            nm = name_field.real_text.strip()
            if not nm:
                toast("أدخل اسم الفرد")
                return
            dlg.dismiss()
            if on_ok:
                on_ok(nm, phone_field.text.strip(), mc_field.text.strip())

        dlg = MDDialog(
            title=ar(title), type="custom", content_cls=box,
            buttons=[
                MDFlatButton(text=ar("إلغاء"), on_release=lambda x: dlg.dismiss()),
                MDRaisedButton(text=ar("حفظ"), on_release=_ok),
            ])
        dlg.open()

    # ---- حوار مبالغ الشعبة (الأصلي + التبرع) ----
    def division_amounts_dialog(self, base=0, donation=0, on_ok=None):
        base_field = MDTextField(
            hint_text=ar("المبلغ الأصلي"), text=_numstr(base),
            input_filter="float", halign="right")
        dona_field = MDTextField(
            hint_text=ar("مبلغ التبرع"), text=_numstr(donation),
            input_filter="float", halign="right")
        box = MDBoxLayout(
            orientation="vertical", spacing=dp(8),
            size_hint_y=None, height=dp(150))
        box.add_widget(base_field)
        box.add_widget(dona_field)

        def _val(f):
            try:
                return float(f.text) if f.text.strip() else 0.0
            except (ValueError, TypeError):
                return 0.0

        def _ok(*a):
            dlg.dismiss()
            if on_ok:
                on_ok(_val(base_field), _val(dona_field))

        dlg = MDDialog(
            title=ar("مبالغ الشعبة"), type="custom", content_cls=box,
            buttons=[
                MDFlatButton(text=ar("إلغاء"), on_release=lambda x: dlg.dismiss()),
                MDRaisedButton(text=ar("حفظ"), on_release=_ok),
            ])
        dlg.open()

    # ---- حوار مبلغ + ملاحظة (مع خيار الخصم) ----
    def amount_dialog(self, title, on_ok=None, show_sign=False):
        deduct = {"v": False}
        amount_field = MDTextField(
            hint_text=ar("المبلغ"), input_filter="float", halign="right")
        note_field = ArabicInput(hint_text=ar("ملاحظة (اختياري)"))
        box = MDBoxLayout(
            orientation="vertical", spacing=dp(8),
            size_hint_y=None, height=dp(150))
        box.add_widget(amount_field)
        box.add_widget(note_field)

        if show_sign:
            row = MDBoxLayout(
                orientation="horizontal", size_hint_y=None,
                height=dp(42), spacing=dp(6))
            cb = MDCheckbox(size_hint=(None, None), size=(dp(40), dp(40)))
            cb.bind(active=lambda inst, val: deduct.update(v=val))
            lbl = MDLabel(text=ar("خصم من الرصيد"), halign="right")
            row.add_widget(cb)
            row.add_widget(lbl)
            box.add_widget(row)
            box.height = dp(200)

        def _ok(*a):
            try:
                amount = float(amount_field.text)
            except (ValueError, TypeError):
                toast("أدخل مبلغًا صحيحًا")
                return
            if show_sign and deduct["v"]:
                amount = -amount
            dlg.dismiss()
            if on_ok:
                on_ok(amount, note_field.real_text)

        dlg = MDDialog(
            title=ar(title), type="custom", content_cls=box,
            buttons=[
                MDFlatButton(text=ar("إلغاء"), on_release=lambda x: dlg.dismiss()),
                MDRaisedButton(text=ar("تطبيق"), on_release=_ok),
            ])
        dlg.open()

    # ---- حوار عرض مسار الملف المُصدَّر ----
    def show_path_dialog(self, path):
        dlg = MDDialog(
            title=ar("تم تصدير الملف"),
            text=ar("حُفظ في:") + "\n" + path,
            buttons=[
                MDFlatButton(
                    text=ar("نسخ المسار"),
                    on_release=lambda x: (self.copy_to_clipboard(path),
                                          dlg.dismiss())),
                MDRaisedButton(
                    text=ar("حسنًا"), md_bg_color=C_PRIMARY,
                    on_release=lambda x: dlg.dismiss()),
            ])
        dlg.open()

    def copy_to_clipboard(self, text):
        try:
            from kivy.core.clipboard import Clipboard
            Clipboard.copy(text)
            toast("تم النسخ")
        except Exception:
            toast("تعذّر النسخ")

    # ---- حوار تأكيد ----
    def confirm_dialog(self, message, on_yes):
        dlg = MDDialog(
            title=ar("تأكيد"), text=ar(message),
            buttons=[
                MDFlatButton(text=ar("إلغاء"), on_release=lambda x: dlg.dismiss()),
                MDRaisedButton(
                    text=ar("نعم"), md_bg_color=(0.8, 0.2, 0.2, 1),
                    on_release=lambda x: (on_yes(), dlg.dismiss())),
            ])
        dlg.open()


if __name__ == "__main__":
    HesabatApp().run()
